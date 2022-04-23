#archivo creado 21-may-2021
#archivo modificado 27-sep-2021
#archivo modificado 7-mar-2022
import openpyxl
#excel_files=['/home/comuser/emt/response/reporte.xlsx']
#file_output='/home/comuser/emt/response/sample_book.xlsx'
file_output='/Users/macbook/Desktop/excel_python/sample_book.xlsx'

from openpyxl import Workbook
# Create a Workbook
wb = Workbook()
ws =  wb.active
ws.title = "Changed Sheet"
sheets= ["TIJUANA", "TECATE", "ROSARITO", "TIJUANA-7330", "TECATE-7330", "ROSARITO-7330", "HUAWEI", "GPON HUAWEI","GPON ISAM"]

for sheet_name in sheets:
	# Create a new Worksheet
	ws1 = wb.create_sheet("")
	# Change sheet name
	ws1.title=sheet_name
	# Create Headers to the current sheet
	ws1.append(["Tecnologia","Equipo","Ip"])

# Create a new Worksheet
#ws1 = wb.create_sheet("")
#ws1.title="TIJUANA"
#ws1.append(["Tecnologia","Equipo","Ip"])

#wb.save(filename = '/home/comuser/emt/response/sample_book.xlsx')
wb.save(file_output)

#wb=openpyxl.load_workbook('/home/comuser/emt/response/sample_book.xlsx')
#abrir archivo como wb donde se van a grabar el filtrado
wb=openpyxl.load_workbook(file_output)

ws=wb.active
ws.cell(row=1, column=1, value=100)
ws.cell(row=2, column=1, value=20)
i =0
text="7302"

#for file in excel_files:
	#workbook=openpyxl.load_workbook(file)
# abrir el archivo principal donde se sacara para filtrado de informacion
#file="/home/comuser/emt/response/reporte.xlsx"
file="/Users/macbook/Desktop/excel_python/reporte.xlsx"
workbook=openpyxl.load_workbook(file)
worksheet=workbook.active
	#worksheet['L6']='=sum(L2:L5)'
	#workbook.save(file)
	# identify the number of occupied rows
	#worksheet.max_row
	# identify the number of occupied rows
	#worksheet.max_column
# sacar el numero de renglones que contiene la columna D del archivo principal	
nrows=len(worksheet['D'])
#print(nrows)

for row in range(1,nrows+1):
	#if str(worksheet.cell(row=row, column=2).value) == text:
  #print row
  found=0
  ciudad=worksheet.cell(row=row, column=2).value
  if ciudad == "TIJUANA" or ciudad == "TECATE" or ciudad == "ROSARITO" :
    if "GPON HUAWEI" in worksheet.cell(row=row, column=4).value :
      find_column=wb["GPON HUAWEI"]
      for j in find_column.iter_rows():
        # loop para buscar si el valor del gpon ya existe en columna 1 de GPON HUAWEI
        if j[1].value == worksheet.cell(row=row, column=22).value :
          found=1
          break
      if found == 0 :    
        wb["GPON HUAWEI"].append([worksheet.cell(row=row, column=4).value,worksheet.cell(row=row, column=22).value,worksheet.cell(row=row, column=6).value])
    if "GPON ISAM" in worksheet.cell(row=row, column=4).value :
      find_column=wb["GPON ISAM"]
      for j in find_column.iter_rows():
        # loop para buscar si el valor del gpon ya existe en columna 1 de GPON ISAM
        if j[1].value == worksheet.cell(row=row, column=22).value :
          found=1
          break
      if found == 0 :    
        wb["GPON ISAM"].append([worksheet.cell(row=row, column=4).value,worksheet.cell(row=row, column=22).value,worksheet.cell(row=row, column=6).value])
    if "IP HUAWEI" in worksheet.cell(row=row, column=4).value :
      wb["HUAWEI"].append([worksheet.cell(row=row, column=4).value,worksheet.cell(row=row, column=5).value,worksheet.cell(row=row, column=6).value])
    if "7330" in worksheet.cell(row=row, column=4).value :
      #print(worksheet.cell(row=row, column=4).value)
      #wb[ciudad+"-7330"].append(["100","200"])
      #Anexar valores tecnologia,nodo, ip en ws de la ciudad que le corresponda el 7330
      wb[ciudad+"-7330"].append([worksheet.cell(row=row, column=4).value,worksheet.cell(row=row, column=5).value,worksheet.cell(row=row, column=6).value])
    if text in worksheet.cell(row=row, column=4).value or "7330" in worksheet.cell(row=row, column=4).value : 
      i = i +1
      nodo=worksheet.cell(row=row, column=5)
      #print(worksheet.cell(row=row, column=5).value)
      print(nodo.value)
		  #ws.cell(row=i, column=1).value=worksheet.cell(row=row, column=5).value
      #guardar valores en el archivo file_output en ws = Changed Sheet
      ws.cell(row=i, column=1).value=nodo.value #asignar valor del nombre del nodo
      ws.cell(row=i, column=2).value=worksheet.cell(row=row, column=6).value  #asignar valor de ip
      ws.cell(row=i, column=3).value=worksheet.cell(row=row, column=4).value  #asignar valor de tecnologia
      #wb['7302'].cell(row=j, column=3).value=worksheet.cell(row=row, column=4).value
      #wb[ciudad].cell(row=j, column=3).value=worksheet.cell(row=row, column=4).value
      wb[ciudad].append([worksheet.cell(row=row, column=4).value,nodo.value,worksheet.cell(row=row, column=6).value]) #agregar a hoja con nombre ciudad valor de tecnologia y nodo 
    #if "7330" in worksheet.cell(row=row, column=4).value : 
     # wb[ciudad+"-7330"].append(["100","200"]) #agregar a hoja con nombre ciudad valor de tecnologia y nodo 
		  #wb.save(filename = 'sample_book.xlsx')
		  #else:
		  #	pass
#wb[ciudad+"-7330"].append(["100","200"]) #agregar a hoja con nombre ciudad valor de tecnologia y nodo 
print("Found %s values that matches the text %s" %(i,text))
#wb.save(filename = 'sample_book.xlsx')
#wb.save('/home/comuser/emt/response/sample_book.xlsx')
wb.save(file_output)