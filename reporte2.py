#codigo creado 29-nov-2020
import openpyxl
excel_files=['/Users/macbook/Desktop/reporte.xlsx']


from openpyxl import Workbook
# Create a Workbook
wb = Workbook()
ws =  wb.active
ws.title = "Changed Sheet"
wb.save(filename = '/Users/macbook/Desktop/sample_book.xlsx')

wb=openpyxl.load_workbook('/Users/macbook/Desktop/sample_book.xlsx')
ws=wb.active
ws.cell(row=1, column=1, value=100)
ws.cell(row=2, column=1, value=20)
i =0
text="TECATE"

#for file in excel_files:
	#workbook=openpyxl.load_workbook(file)
file="/Users/macbook/Desktop/reporte.xlsx"
workbook=openpyxl.load_workbook(file)
worksheet=workbook.active
	#worksheet['L6']='=sum(L2:L5)'
	#workbook.save(file)
	# identify the number of occupied rows
	#worksheet.max_row
	# identify the number of occupied rows
	#worksheet.max_column
nrows=len(worksheet['B'])
print(nrows)
for row in range(1,nrows+1):
	if str(worksheet.cell(row=row, column=2).value) == text:
		i = i +1
		c=worksheet.cell(row=row, column=5)
		#print(worksheet.cell(row=row, column=5).value)
		print(c.value)
		#ws.cell(row=row, column=1).value=worksheet.cell(row=row, column=5).value
		ws.cell(row=row, column=1).value=c.value
		#wb.save(filename = 'sample_book.xlsx')
		#else:
		#	pass
print("Found %s Email Names that matches the %s domain" %(i,text))
#wb.save(filename = 'sample_book.xlsx')
wb.save('/Users/macbook/Desktop/sample_book.xlsx')
