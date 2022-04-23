
# importing openpyxl module 
#codigo creado 29-nov-2020
import openpyxl as xl; 

from openpyxl import Workbook
#Create a Workbook
wb = Workbook()
ws =  wb.active
#Change the sheet title
ws.title = "Changed Sheet"
#Save the new workbook as
wb.save(filename = '/Users/macbook/Desktop/sample_book.xlsx')
  
# opening the source excel file 
filename ="/Users/macbook/Desktop/reporte.xlsx"
wb1 = xl.load_workbook(filename) 
ws1 = wb1.worksheets[0] 
  
# opening the destination excel file  
filename1 ="/Users/macbook/Desktop/sample_book.xlsx"
wb2 = xl.load_workbook(filename1) 
ws2 = wb2.active 
  
# calculate total number of rows and  
# columns in source excel file 
mr = ws1.max_row 
mc = ws1.max_column 
col=0
# copying the cell values from source  
# excel file to destination excel file 
for i in range (1, mr + 1): 
	col=0
    	for j in range (1, mc + 1): 
    		if j == 2 or j == 3 or j == 4 or j == 5 or j == 6 or j == 22:
    			col=col+1
        		# reading cell value from source excel file 
        		c = ws1.cell(row = i, column = j) 
  
        		# writing the read value to destination excel file 
        		ws2.cell(row = i, column = col).value = c.value 
  
# saving the destination excel file 
wb2.save(str(filename1)) 
